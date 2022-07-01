# -*- coding: utf-8 -*-
# Purpose: 解析excel数据表到lua

import os
import sys
import string
import xlrd
import json
import openpyxl
import warnings

warnings.simplefilter("ignore")

_errorDes = ["数据错误：", "文件名", "行数", "列数"]
_isError = False

global luaPath
global excelPath
global configname

global isXls

# 按类型解析数据
dataType = ["int", "float", "string", "bool", "table"]


# 查找所有 xls
def findAllFile(callback):
    fileList = os.listdir(excelPath)
    for f in fileList:

        # 检查是否有错误
        if _isError:
            break

        filePath = os.path.join(excelPath, f)

        if f[0] == "." or f.find(".svn") > 0 or f.find(".DS_Store") > 0 or f.startswith("~$"):
            continue

        if os.path.isdir(filePath):
            findAllFile(filePath, callback)
        else:
            if f.endswith(".xls") or f.endswith(".xlsx"):
                callback(filePath, f)


# 解析 excel的type类型
# ctype类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
def parseTitleType(rowData, col):
    tableKey = []
    tableType = []
    for c in range(col):
        keyValue = None
        typeValue = None
        if isXls:
            keyValue = rowData.cell(2, c).value
            typeValue = rowData.cell(1, c).value
        else:
            keyValue = rowData.cell(row=3, column=c + 1).value
            typeValue = rowData.cell(row=2, column=c + 1).value
        if typeValue is None:
            break
        tmp = str(typeValue).split('|')
        # 这里判断只导出客户端的
        if len(tmp) == 2:
            if tmp[1] == 'client':
                typeValue = tmp[0]
            else:
                typeValue = None
        tableKey.append(keyValue)
        tableType.append(typeValue)
    return tableKey, tableType


# 获取表格数据
def getDataByExcel(pTabData, pRow, tblKey):
    result = []
    for row in range(pRow):
        coltab = []
        for col in range(len(tblKey)):
            tmp = None
            global isXls
            if isXls:
                tmp = pTabData.cell(row + 3, col).value
            else:
                tmp = pTabData.cell(row=row + 4, column=col + 1).value
            if tmp is None:
                tmp = 'nil'
            coltab.append(tmp)
        result.append(coltab)
    return result


# ******************************************************* 格式化输出数据
# 文件介绍
luaTblName = "data"
_luaDes = '''-- Filename: %s.lua
-- methods: X.getitem(id)
-- Function: no description.\n
'''
_luaTblData = "local {} ".format(luaTblName) + "= {\n%s}\n\n"
_luaTblItemData = "\t[%s] = {%s },\n"
_luaDeclare = "local %s = {}\n"
_luaLookupFunc = '''function %s.get(id)\n\tif not id or type(id) ~= 'number' then return nil end\n'''
_lualookupFuncReturn = '''\treturn {}[id] end\n'''.format(luaTblName)
_luaGetAllFunc = '''function %s.getall() return {} end\n'''.format(luaTblName)
_luaReturn = '''return %s'''


# *******************************************************
# 将excel行数据整理成lua的table
def parseData(tblData, titleTypes, titleKeys):
    result = ""
    for row in range(len(tblData)):
        item = []
        for col in range(len(titleTypes)):
            if titleTypes[col] is None or len(str(titleTypes[col])) == 0 or titleKeys[col] is None or len(
                    str(titleKeys[col])) == 0:
                continue
            tmp = tblData[row][col]
            empty = False
            if tmp is None or tmp == '':
                tmp = 'nil'
                empty = True
            if titleTypes[col] == dataType[0]:
                if not empty:
                    tmp = int(tmp)
            elif titleTypes[col] == dataType[1]:
                if not empty:
                    tmp = float(tmp)
            elif titleTypes[col] == dataType[2]:
                if not empty:
                    tmp = "\"%s\"" % (str(tmp).replace('\"', '\\\"'))
            elif titleTypes[col] == dataType[3]:
                if tmp == 'false' or empty:
                    tmp = 'false'
            elif titleTypes[col] == dataType[4]:
                if not empty:
                    rrr = []
                    vtmp = tmp.split(',')
                    if len(vtmp) > 0:
                        for i in range(len(vtmp)):
                            vvtmp = vtmp[i].split('=')
                            if len(vvtmp) == 2:
                                fff = vvtmp[0]
                                sss = ("\"%s\"" % (vvtmp[1].replace('\"', '\\\"')))
                                if sss == '"true"':
                                    sss = 'true'
                                elif sss == '"false"':
                                    sss = 'false'
                                if fff.isdigit():
                                    rrr.append('[' + str(fff) + '] = ' + sss)
                                else:
                                    rrr.append(str(fff) + ' = ' + sss)
                            else:
                                rrr.append(str(vtmp[i]))
                        tmp = '{%s}' % ','.join(rrr)
                    else:
                        tmp = '{%s}' % (tmp.replace('\"', '\\\"'))
            item.append((' %s = ' % titleKeys[col]) + str(tmp))
        itemData = _luaTblItemData % ((len(titleKeys) > 0 and titleKeys[0] == 'id' and titleTypes[0] == 'int')
                                      and int(tblData[row][0]) or (row + 1), ','.join(item))
        result = result + itemData
    return result


def parseAnnotation(titleTypes, titleKeys):
    result = "\t---@class %s\n" % configname
    for i in range(len(titleKeys)):
        if titleKeys[i] != None and len(str(titleKeys[i])) > 0:
            type = titleTypes[i]
            if titleTypes[i] == dataType[0] or titleTypes[i] == dataType[1]:
                type = "number"
            result += "\t---@field {} {}\n".format(titleKeys[i], type)
    return result


# 构造输出lua
def getLuaText(title, tabData, fileName, titleTypes, titleKeys):
    global configname
    configname = title + "Config"
    modulename = "conf_%s" % (title.capitalize())
    # 文件介绍
    luaDes = _luaDes % modulename
    # 文件数据内容
    luaTblData = ("---@type %s[]\n" % configname) + _luaTblData % (parseData(tabData, titleTypes, titleKeys))
    # 声明lua表
    luaDeclare = ("---@class %s\n" % modulename) + _luaDeclare % modulename
    # 查询函数
    luaLookupFuncReturn = parseAnnotation(titleTypes, titleKeys) + _lualookupFuncReturn
    luaLookupFunc = ('---@return %s\n' % configname) + _luaLookupFunc % modulename + luaLookupFuncReturn
    # 获取所有数据
    luaGetAllFunc = _luaGetAllFunc % modulename
    # 文件返回
    luaReturn = _luaReturn % modulename
    content = luaDes + luaTblData + luaDeclare + luaLookupFunc + luaGetAllFunc + luaReturn
    return modulename, content


# 写文件
def writeFile(filePath, fileData):
    f = open(filePath, "w", encoding="utf-8")
    f.write(fileData)
    f.close()


# 写文件
def writeToLua(filePath, fileName, fileData):
    global luaPath
    global excelPath
    fDirPath = os.path.dirname(filePath)
    fDirPath = fDirPath.replace(excelPath, luaPath)
    if not os.path.exists(fDirPath):
        os.mkdir(fDirPath)
    filePath = os.path.join(fDirPath, "%s.lua" % (fileName))
    writeFile(filePath, fileData)


# 拆分excel
def parseExcel(filePath, fileName):
    global isXls
    if fileName.endswith(".xls") or fileName.endswith(".xlsx"):
        isXls = True
        # 读取数据
        excel = xlrd.open_workbook(filePath)
        # 获取workbook中所有的表格
        sheets = excel.sheet_names()
        # 循环遍历所有sheet
        sheetslen = len(sheets)
        for i in range(sheetslen):
            sheet = excel.sheet_names()[i]
            print('parsing excel {},sheet:{}'.format(filePath, sheet))
            pTabData = excel.sheet_by_index(i)
            if pTabData.nrows > 3 and pTabData.ncols > 0:
                tabTitle = parseTitleType(pTabData, pTabData.ncols)
                data = getDataByExcel(pTabData, pTabData.nrows - 3, tabTitle[0])
                # title = sheet.split('|')
                configname, suffix = os.path.splitext(fileName)
                if sheetslen > 1:
                    configname += str(i)
                filedata = getLuaText(configname, data, fileName, tabTitle[1], tabTitle[0])
                writeToLua(filePath, filedata[0], filedata[1])
            else:
                print("%s表无内容！！！" % filePath)
    else:
        isXls = False
        excel = openpyxl.load_workbook(filePath)
        # sheets = excel.get_sheet_names()
        sheets = excel.sheetnames
        # 循环遍历所有sheet
        sheetslen = len(sheets)
        for i in range(sheetslen):
            sheet = excel[sheets[i]]
            if sheet.max_row > 3 and sheet.max_column > 0:
                tabTitle = parseTitleType(sheet, sheet.max_column)
                data = getDataByExcel(sheet, sheet.max_row - 3, tabTitle[0])
                configname, suffix = os.path.splitext(fileName)
                if sheetslen > 1:
                    configname += str(i)
                filedata = getLuaText(configname, data, fileName, tabTitle[1], tabTitle[0])
                # 写入文件
                writeToLua(filePath, filedata[0], filedata[1])
            else:
                print("%s表无内容！！！" % filePath)


# 1.excel 文件夹路径 2.lua文件路径
if __name__ == "__main__":
    global excelPath
    global luaPath
    luaPath = os.getcwd() + "\\lua_configs"
    excelPath = os.getcwd() + "\\excel_configs"
    if not os.path.exists(excelPath):
        print('excelPath: %s 路径不存在' % excelPath)
        sys.exit(1)
    # 检查并创建目录
    if not os.path.exists(luaPath):
        os.makedirs(luaPath)
    findAllFile(parseExcel)
    print('parse excel success!')
else:
    print('参数不正确')
    sys.exit(1)
